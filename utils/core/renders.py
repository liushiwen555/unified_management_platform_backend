from typing import Mapping, List
from tempfile import NamedTemporaryFile

from django.conf import settings
from rest_framework.renderers import BaseRenderer
from openpyxl import Workbook
from docxtpl import DocxTemplate


class XLSXRenderer(BaseRenderer):
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


class DOCXRenderer(BaseRenderer):
    media_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    format = 'docx'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


class ExportXLSXRenderer(XLSXRenderer):
    """
    Renderer which serializes to xlsx.
    """
    # https://developer.mozilla.org/en-US/docs/Web/HTTP/Basics_of_HTTP/MIME_types/Common_types
    # https://docs.microsoft.com/en-us/microsoft-365/compliance/supported-filetypes-datainvestigations?view=o365-worldwide
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        """
        Render `data` into xlsx, returning a bytestring.
        """
        try:
            data = self._change_data_to_save(data)
            wb = self.export_device(data)
            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                stream = tmp.read()

            return super().render(stream, accepted_media_type, renderer_context)
        except AttributeError:
            return super().render(data, accepted_media_type, renderer_context)

    def export_device(self, data):
        workbook = Workbook()
        sheet = workbook.create_sheet('资产导出列表', 0)
        head = [
            '资产ID',
            '添加时间',
            '资产名称',
            '资产类别',
            '资产类型',
            '厂商',
            '型号',
            '版本',
            '资产IP',
            '资产MAC',
            '负责人',
            '资产位置',
            '重要程度',
            '备注',
        ]

        for h in range(len(head)):
            sheet.cell(1, h + 1, head[h])

        i = 2
        for product in data:
            sheet.cell(i, 1, product['id'])
            sheet.cell(i, 2, product['created_at'])
            sheet.cell(i, 3, product['name'])
            sheet.cell(i, 4, product['category'])
            sheet.cell(i, 5, product['type'])
            sheet.cell(i, 6, product['brand'])
            sheet.cell(i, 7, product['hardware'])
            sheet.cell(i, 8, product['version'])
            sheet.cell(i, 9, product['ip'])
            sheet.cell(i, 10, product['mac'])
            sheet.cell(i, 11, product['responsible_user'])
            sheet.cell(i, 12, product['location'])
            sheet.cell(i, 13, product['value']),
            sheet.cell(i, 14, product['description'])
            i += 1

        return workbook

    def _change_data_to_save(self, data):
        ori_data = data
        for d in ori_data:
            if d.get('category') == 1:
                d['category'] = '安全设备'

            if d.get('category') == 2:
                d['category'] = '网络设备'

            if d.get('category') == 3:
                d['category'] = '主机设备'

            if d.get('category') == 4:
                d['category'] = '工控设备'

        return ori_data


class ExportDOCXRenderer(DOCXRenderer):
    def render(self, data, accepted_media_type=None, renderer_context=None):
        try:
            data = self.save_data_to_word(data)
            with NamedTemporaryFile() as tmp:
                data.save(tmp.name)
                tmp.seek(0)
                stream = tmp.read()
            return super().render(stream, accepted_media_type, renderer_context)
        except AttributeError:
            return super().render(data, accepted_media_type, renderer_context)

    def save_data_to_word(self, data):
        report_path = settings.MEDIA_ROOT + 'report.docx'
        tpl = DocxTemplate(report_path)
        tpl.render(data)

        return tpl
